from django.shortcuts import render, redirect, HttpResponse, get_object_or_404
from .forms import SiswaForm, GuruForm, TindakanForm
from .models import Siswa, Guru, Tindakan, Pelanggaran
import barcode
from barcode.writer import ImageWriter
import qrcode
import base64
from io import BytesIO
from django.urls import reverse
import requests
from django.conf import settings
import os
from .forms import PelanggaranForm
from .models import Pelanggaran, SiswaPelanggaran

from django.shortcuts import render, redirect
from .forms import GuruForm

from rest_framework import viewsets
from .serializers import SiswaSerializer

from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password
from django.db import models
from django.db.models import F

from django.contrib.auth import logout

def dashboard(request):
    return render(request, 'dashboard.html')

import os
from django.conf import settings
from django.shortcuts import render, redirect
import qrcode
from io import BytesIO
from .forms import SiswaForm

def create_siswa(request):
    if request.method == 'POST':
        form = SiswaForm(request.POST)
        if form.is_valid():
            siswa = form.save(commit=False)

            qr_code_url = request.build_absolute_uri(reverse('barcode_scanner', args=[siswa.nomor_induk]))
            qr_code = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=4)
            qr_code.add_data(qr_code_url)
            qr_code.make(fit=True)
            qr_code_image = qr_code.make_image(fill_color="black", back_color="white")

            qr_code_buffer = BytesIO()
            qr_code_image.save(qr_code_buffer, format='PNG')

            qr_code_directory = os.path.join('media', 'siswa')
            os.makedirs(qr_code_directory, exist_ok=True)

            qr_code_filepath = os.path.join(qr_code_directory, f'qr_code_{siswa.nomor_induk}.png')

            with open(qr_code_filepath, 'wb') as f:
                f.write(qr_code_buffer.getvalue())

            siswa.qr_code = qr_code_filepath
            siswa.save()

            return redirect('siswa_list')
    else:
        form = SiswaForm()

    return render(request, 'siswa/create.html', {'form': form})

def siswa_list(request):
    siswa = Siswa.objects.all()
    for s in siswa:
        qr_code = qrcode.make(s.nomor_induk)
        buffered = BytesIO()
        qr_code.save(buffered, format='PNG')
        qr_code_base64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
        s.qr_code = qr_code_base64
    return render(request, 'siswa/list.html', {'siswa': siswa})

def update_siswa(request, pk):
    siswa = Siswa.objects.get(pk=pk)
    if request.method == 'POST':
        form = SiswaForm(request.POST, instance=siswa)
        if form.is_valid():
            form.save()
            return redirect('siswa_list')
    else:
        form = SiswaForm(instance=siswa)
    return render(request, 'siswa/update.html', {'form': form, 'siswa': siswa})

def delete_siswa(request, pk):
    siswa = Siswa.objects.get(pk=pk)
    siswa.delete()
    return redirect('siswa_list')

def generate_barcode(data):
    barcode_class = barcode.get_barcode_class('code128')
    barcode_image = barcode_class(data, writer=ImageWriter())
    buffer = BytesIO()
    barcode_image.write(buffer)
    buffer.seek(0)
    return buffer

def barcode_scanner(request, nomor_induk):
    siswa = get_object_or_404(Siswa, nomor_induk=nomor_induk)
    pelanggaran = Pelanggaran.objects.all()
    return render(request, 'siswa/TambahPelanggaranSiswa.html', {'siswa': [siswa], 'pelanggaran_list': pelanggaran})

def add_pelanggaran_siswa(request):
    if request.method == 'POST':
        barcode_data = request.POST.get('id_siswa')
        nama_pelanggaran = request.POST.get('nama_pelanggaran')
        
        siswa = Siswa.objects.get(id=barcode_data)
        pelanggaran = Pelanggaran.objects.get(id_pelanggaran=nama_pelanggaran)
        
        pelanggaran_siswa = SiswaPelanggaran(pelanggaran=pelanggaran, siswa=siswa)
        pelanggaran_siswa.save()
        
    return redirect('rekap_laporan')


def guru_list(request):
    gurus = Guru.objects.all()
    return render(request, 'guru/list.html', {'guru': gurus})

from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password

def guru_create(request):
    if request.method == 'POST':
        form = GuruForm(request.POST)
        if form.is_valid():
            guru = form.save(commit=False)
            username = guru.id_guru
            password = make_password(guru.id_guru)
            user = User.objects.create(username=username, password=password)
            guru.user = user
            guru.save()
            return redirect('guru_list')
    else:
        form = GuruForm()
    return render(request, 'guru/create.html', {'form': form})

def guru_update(request, guru_id):
    guru = get_object_or_404(Guru, id=guru_id)
    if request.method == 'POST':
        form = GuruForm(request.POST, instance=guru)
        if form.is_valid():
            form.save()
            return redirect('guru_list')
    else:
        form = GuruForm(instance=guru)
    return render(request, 'guru/update.html', {'form': form, 'guru': guru})

def guru_delete(request, guru_id):
    guru = get_object_or_404(Guru, id=guru_id)
    guru.delete()
    return redirect('guru_list')

# Data Pelanggaran
def pelanggaran_list(request):
    pelanggaran = Pelanggaran.objects.all()
    return render(request, 'pelanggaran/list.html', {'pelanggaran': pelanggaran})

def pelanggaran_create(request):
    if request.method == 'POST':
        form = PelanggaranForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('pelanggaran_list')
    else:
        form = PelanggaranForm()
    return render(request, 'pelanggaran/create.html', {'form': form})

def pelanggaran_update(request, pelanggaran_id):
    pelanggaran = Pelanggaran.objects.get(id_pelanggaran=pelanggaran_id)
    if request.method == 'POST':
        form = PelanggaranForm(request.POST, instance=pelanggaran)
        if form.is_valid():
            form.save()
            return redirect('pelanggaran_list')
    else:
        form = PelanggaranForm(instance=pelanggaran)
    return render(request, 'pelanggaran/update.html', {'form': form, 'pelanggaran_id': pelanggaran_id})

def pelanggaran_delete(request, pelanggaran_id):
    pelanggaran = get_object_or_404(Pelanggaran, id_pelanggaran=pelanggaran_id)
    pelanggaran.delete()
    return redirect('pelanggaran_list')

class SiswaViewSet(viewsets.ModelViewSet):
    queryset = Siswa.objects.all()
    serializer_class = SiswaSerializer

def list_siswa(request):
    response = requests.get('http://localhost:8000/api/siswa/')  # Assuming the API URL is 'http://localhost:8000/api/siswa/'
    if response.status_code == 200:
        data = response.json()
        siswa_list = data.get('results', [])  # Retrieve the 'results' key from the response data
        return render(request, 'siswa_list.html', {'siswa_list': siswa_list})

    return render(request, 'error.html', {'message': 'Failed to retrieve data from the API.'})

def generate_qrcode(request, data):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)

    # Create an image from the QR code data
    image = qr.make_image(fill_color="black", back_color="white")

    # Create a response object and return the image
    response = HttpResponse(content_type="image/png")
    image.save(response, "PNG")
    return response

from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from .forms import LoginForm

def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                request.session['logged_in'] = True
                return redirect('dashboard')  # Replace 'home' with your desired URL name or path
            else:
                form.add_error(None, 'Invalid username or password')
    else:
        form = LoginForm()
    
    return render(request, 'login.html', {'form': form})

from .models import Tindakan
from .forms import TindakanForm

def tindakan_list(request):
    tindakan_list = Tindakan.objects.all()
    return render(request, 'tindakan/list.html', {'tindakan_list': tindakan_list})

def tindakan_create(request):
    if request.method == 'POST':
        form = TindakanForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('tindakan_list')
    else:
        form = TindakanForm()
    return render(request, 'tindakan/create.html', {'form': form})

def tindakan_update(request, tindakan_id):
    tindakan = Tindakan.objects.get(id_tindakan=tindakan_id)
    if request.method == 'POST':
        form = TindakanForm(request.POST, instance=tindakan)
        if form.is_valid():
            form.save()
            return redirect('tindakan_list')
    else:
        form = TindakanForm(instance=tindakan)
    return render(request, 'tindakan/update.html', {'form': form, 'tindakan_id': tindakan_id})

def tindakan_delete(request, tindakan_id):
    tindakan = get_object_or_404(Tindakan, id_tindakan=tindakan_id)
    tindakan.delete()
    return redirect('tindakan_list')

from .models import Siswa, Pelanggaran, Tindakan, SiswaPelanggaran

def rekap_laporan(request):
    siswa_pelanggaran_data = SiswaPelanggaran.objects.prefetch_related('siswa', 'pelanggaran').all()

    for siswa_pelanggaran in siswa_pelanggaran_data:
        pelanggaran = siswa_pelanggaran.pelanggaran
        tindakan = Tindakan.objects.filter(range_point__gte=pelanggaran.point_pelanggaran).order_by('range_point').first()
        siswa_pelanggaran.sanksi = tindakan.tindakan if tindakan else 'Tidak ada sanksi' # Jika tidak ada tindakan, tampilkan pesan "Tidak ada sanksi"

    return render(request, 'rekap_laporan.html', {'siswa_pelanggaran_data': siswa_pelanggaran_data})

def logout_view(request):
    logout(request)
    request.session.clear()
    return redirect('login')

def sync_siswa(request):
    response = requests.get('http://127.0.0.1:8000/api/siswa/?format=json')  # Ubah URL sesuai dengan API Anda
    if response.status_code == 200:
        data = response.json()
        siswa_list = data  # Assuming the API response is a list of student objects directly
        return render(request, 'siswa_list.html', {'siswa_list': siswa_list})

    return render(request, 'error.html', {'message': 'Failed to retrieve data from the API.'})

def create_pelanggaran_manual(request):
    siswa = Siswa.objects.all()
    return render(request, 'siswa/TambahPelanggaranManualSiswa.html', {'siswa': siswa})

def cetak_semua_laporan(request):
    siswa_pelanggaran_data = SiswaPelanggaran.objects.select_related('siswa', 'pelanggaran').all()
    for siswa_pelanggaran in siswa_pelanggaran_data:
        pelanggaran = siswa_pelanggaran.pelanggaran
        tindakan = Tindakan.objects.filter(range_point__gte=pelanggaran.point_pelanggaran).order_by('range_point').first()
        siswa_pelanggaran.sanksi = tindakan.tindakan if tindakan else 'Tidak ada sanksi' # Jika tidak ada tindakan, tampilkan pesan "Tidak ada sanksi"

    return render(request, 'cetak_siswa.html', {'siswa_pelanggaran_data': siswa_pelanggaran_data})


import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from io import BytesIO
from django.http import HttpResponse
import json

def export_to_excel(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        table_data = data.get('data', [])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=data_pelanggaran_siswa.xlsx'

        wb = openpyxl.Workbook()
        ws = wb.active

        # Add table headers
        headers = ["Nomor Induk", "Nama Siswa", "Nama Pelanggaran", "Point"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header

        for row_num, row_data in enumerate(table_data, 2):
            for col_num, cell_data in enumerate(row_data, 1):
                col_letter = get_column_letter(col_num)
                ws[f"{col_letter}{row_num}"] = cell_data

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Send the file in the response
        response.write(excel_file.read())
        return response

from django.shortcuts import render, redirect
from .models import Siswa
from .forms import SiswaForm
import openpyxl

def import_siswa_from_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']

        if excel_file.name.endswith('.xlsx'):
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                nomor_induk, nama_lengkap, tempat_lahir, tanggal_lahir, jenis_kelamin, kelas, rombel = row
                Siswa.objects.create(
                    nomor_induk=nomor_induk,
                    nama_lengkap=nama_lengkap,
                    tempat_lahir=tempat_lahir,
                    tanggal_lahir=tanggal_lahir,
                    jenis_kelamin=jenis_kelamin,
                    kelas=kelas,
                    rombel=rombel
                )

            return redirect('siswa_list')
    return render(request, 'siswa/import_excel.html')

